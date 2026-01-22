from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count, Sum, F, Case, When, IntegerField, Avg, Q
from django.utils import timezone
from calendar import month_name
from datetime import timedelta, datetime
from galeri.models import Arac, personel
from galeri.decorators import rapor_required


def _get_aylik_satislar(satilan_araclar, baslangic_yil, bitis_yil, baslangic_ay=0, bitis_ay=0):
    """Aylık satışları Django ORM ile verimli şekilde hesapla"""
    aylik_satislar = {}
    aylik_gelir = {}
    aylik_kar = {}
    
    # Ay aralığını belirle
    baslangic_ay_range = baslangic_ay if baslangic_ay > 0 else 1
    bitis_ay_range = bitis_ay if bitis_ay > 0 else 12
    
    for yil in range(baslangic_yil, bitis_yil + 1):
        # Yıl başlangıç ve bitiş ayını belirle
        if yil == baslangic_yil:
            yil_baslangic_ay = baslangic_ay_range
        else:
            yil_baslangic_ay = 1
        
        if yil == bitis_yil:
            yil_bitis_ay = bitis_ay_range
        else:
            yil_bitis_ay = 12
        
        for ay in range(yil_baslangic_ay, yil_bitis_ay + 1):
            ay_adi = f"{month_name[ay]} {yil}"
            
            # Django ORM ile filtreleme - çok daha hızlı
            ay_satislari = satilan_araclar.filter(
                Q(satis_tarihi__year=yil, satis_tarihi__month=ay) |
                Q(satis_tarihi__isnull=True, ilan_tarihi__year=yil, ilan_tarihi__month=ay)
            )
            
            aylik_satislar[ay_adi] = ay_satislari.count()
            
            # Aylık gelir
            gelir = ay_satislari.aggregate(
                toplam=Sum(Case(
                    When(satis_fiyat__isnull=False, then=F('satis_fiyat')),
                    default=F('ilan_fiyat'),
                    output_field=IntegerField()
                ))
            )['toplam'] or 0
            aylik_gelir[ay_adi] = gelir
            
            # Aylık kar - sadece kar_zarar null olmayan kayıtları say
            kar = ay_satislari.filter(kar_zarar__isnull=False).aggregate(
                toplam=Sum('kar_zarar')
            )['toplam']
            aylik_kar[ay_adi] = kar if kar is not None else 0
    
    return aylik_satislar, aylik_gelir, aylik_kar


def _get_marka_aylik_satislar(satilan_araclar, baslangic_yil, bitis_yil):
    """Markalara göre aylık satışları hesapla"""
    marka_aylik_satislar = {}
    markalar = satilan_araclar.values_list('marka__ad', flat=True).distinct()
    
    for marka in markalar:
        marka_aylik_satislar[marka] = {}
        marka_araclari = satilan_araclar.filter(marka__ad=marka)
        
        for yil in range(baslangic_yil, bitis_yil + 1):
            for ay in range(1, 13):
                ay_adi = f"{month_name[ay]} {yil}"
                
                count = marka_araclari.filter(
                    Q(satis_tarihi__year=yil, satis_tarihi__month=ay) |
                    Q(satis_tarihi__isnull=True, ilan_tarihi__year=yil, ilan_tarihi__month=ay)
                ).count()
                
                marka_aylik_satislar[marka][ay_adi] = count
    
    return marka_aylik_satislar


def _get_toplam_gelir(satilan_araclar):
    """Toplam geliri hesapla"""
    return satilan_araclar.aggregate(
        toplam=Sum(Case(
            When(satis_fiyat__isnull=False, then=F('satis_fiyat')),
            default=F('ilan_fiyat'),
            output_field=IntegerField()
        ))
    )['toplam'] or 0


def _get_toplam_kar(satilan_araclar):
    """Toplam karı hesapla"""
    return satilan_araclar.aggregate(toplam=Sum('kar_zarar'))['toplam'] or 0



def _get_personel_performans(satilan_araclar):
    """Personel bazında satış ve kar performansını hesapla"""
    
    personel_performans = []
    
    # Sadece bu satılan araçlarda yer alan personelleri al
    personel_ids = satilan_araclar.filter(
        personel__isnull=False
    ).values_list('personel_id', flat=True).distinct()
    
    personeller = personel.objects.filter(id__in=personel_ids)
    
    for pers in personeller:
        # Bu personelin sattığı araçlar (zaten filtrelenmiş satilan_araclar içinden)
        personel_satislari = satilan_araclar.filter(personel=pers)
        satis_adedi = personel_satislari.count()
        
        # Satış yoksa atla
        if satis_adedi == 0:
            continue
        
        # Toplam kar (sadece kar_zarar null olmayan kayıtlar)
        toplam_kar = personel_satislari.filter(
            kar_zarar__isnull=False
        ).aggregate(toplam=Sum('kar_zarar'))['toplam'] or 0
        
        personel_performans.append({
            'personel': f"{pers.ad} {pers.soyad}",
            'satis_adedi': satis_adedi,
            'toplam_kar': toplam_kar
        })
    
    # Satış adedine göre sırala
    return sorted(personel_performans, key=lambda x: x['satis_adedi'], reverse=True)


@rapor_required
def raporlar(request):
    """Satış raporları ana sayfası - Kapsamlı analiz"""
    # Tarih aralığı parametreleri
    baslangic_yil = request.GET.get('baslangic_yil', timezone.now().year)
    bitis_yil = request.GET.get('bitis_yil', timezone.now().year)
    baslangic_ay = request.GET.get('baslangic_ay', 0)
    bitis_ay = request.GET.get('bitis_ay', 0)
    
    try:
        baslangic_yil = int(baslangic_yil)
        bitis_yil = int(bitis_yil)
        baslangic_ay = int(baslangic_ay)
        bitis_ay = int(bitis_ay)
    except (ValueError, TypeError):
        baslangic_yil = timezone.now().year
        bitis_yil = timezone.now().year
        baslangic_ay = 0
        bitis_ay = 0
    
    # Ay değerlerini kontrol et (0-12 arası)
    if baslangic_ay < 0 or baslangic_ay > 12:
        baslangic_ay = 0
    if bitis_ay < 0 or bitis_ay > 12:
        bitis_ay = 0
    
    # Yıl listesi oluştur (2020'den mevcut yıla kadar)
    mevcut_yil = timezone.now().year
    yil_listesi = list(range(2020, mevcut_yil + 1))
    yil_listesi.reverse()  # En yeni yıl önce gelsin
    
    # ============ TEMEL İSTATİSTİKLER ============
    # Satıldı olarak işaretlenmiş araçları al
    satilan_araclar = Arac.objects.filter(satildi_mi=True)
    
    # Tarih filtresi uygula
    if baslangic_ay > 0:
        # Başlangıç ayı belirtilmişse
        satilan_araclar = satilan_araclar.filter(
            Q(satis_tarihi__year__gte=baslangic_yil, satis_tarihi__month__gte=baslangic_ay) |
            Q(satis_tarihi__year__gt=baslangic_yil) |
            Q(satis_tarihi__isnull=True, ilan_tarihi__year__gte=baslangic_yil, ilan_tarihi__month__gte=baslangic_ay) |
            Q(satis_tarihi__isnull=True, ilan_tarihi__year__gt=baslangic_yil)
        )
    else:
        # Sadece yıl filtresi
        satilan_araclar = satilan_araclar.filter(
            Q(satis_tarihi__year__gte=baslangic_yil) |
            Q(satis_tarihi__isnull=True, ilan_tarihi__year__gte=baslangic_yil)
        )
    
    if bitis_ay > 0:
        # Bitiş ayı belirtilmişse
        satilan_araclar = satilan_araclar.filter(
            Q(satis_tarihi__year__lte=bitis_yil, satis_tarihi__month__lte=bitis_ay) |
            Q(satis_tarihi__year__lt=bitis_yil) |
            Q(satis_tarihi__isnull=True, ilan_tarihi__year__lte=bitis_yil, ilan_tarihi__month__lte=bitis_ay) |
            Q(satis_tarihi__isnull=True, ilan_tarihi__year__lt=bitis_yil)
        )
    else:
        # Sadece yıl filtresi
        satilan_araclar = satilan_araclar.filter(
            Q(satis_tarihi__year__lte=bitis_yil) |
            Q(satis_tarihi__isnull=True, ilan_tarihi__year__lte=bitis_yil)
        )
    
    # Aylık satışlar (optimize edilmiş)
    aylik_satislar, aylik_gelir, aylik_kar = _get_aylik_satislar(
        satilan_araclar, baslangic_yil, bitis_yil, baslangic_ay, bitis_ay
    )
    
    # Markalara göre aylık satışlar
    marka_aylik_satislar = _get_marka_aylik_satislar(
        satilan_araclar, baslangic_yil, bitis_yil
    )
    
    # ============ TOPLAM İSTATİSTİKLER ============
    toplam_satilan = satilan_araclar.count()
    toplam_gelir = _get_toplam_gelir(satilan_araclar)
    toplam_kar = _get_toplam_kar(satilan_araclar)
    
    # Ortalama satış fiyatı
    ortalama_satis_fiyati = toplam_gelir / toplam_satilan if toplam_satilan > 0 else 0
    
    # Ortalama kar
    ortalama_kar = toplam_kar / toplam_satilan if toplam_satilan > 0 else 0
    
    # ============ KARLILIK ANALİZİ ============
    # En karlı araçlar (top 10)
    en_karli_araclar = satilan_araclar.filter(
        kar_zarar__isnull=False
    ).order_by('-kar_zarar')[:10]
    
    # En az karlı araçlar (bottom 10)
    en_az_karli_araclar = satilan_araclar.filter(
        kar_zarar__isnull=False
    ).order_by('kar_zarar')[:10]
    
    # Marka bazında karlılık
    marka_karlilik = []
    markalar = satilan_araclar.values_list('marka__ad', flat=True).distinct()
    
    # Marka bazında kar marjı yüzdesi (grafik için)
    marka_kar_marji = []
    
    for marka in markalar:
        marka_araclari = satilan_araclar.filter(marka__ad=marka)
        toplam_kar_marka = _get_toplam_kar(marka_araclari)
        toplam_gelir_marka = _get_toplam_gelir(marka_araclari)
        satis_sayisi = marka_araclari.count()
        ortalama_kar_marka = toplam_kar_marka / satis_sayisi if satis_sayisi > 0 else 0
        
        # Kar marjı yüzdesi = (kar / gelir) * 100
        # Sadece gelir > 0 ve kar_zarar null olmayan markaları ekle
        if toplam_gelir_marka > 0 and toplam_kar_marka is not None:
            kar_marji_yuzdesi = (toplam_kar_marka / toplam_gelir_marka * 100)
            # Negatif kar marjı da gösterilebilir (zarar durumu)
            marka_kar_marji.append({
                'marka': marka,
                'kar_marji_yuzdesi': round(kar_marji_yuzdesi, 2),
                'toplam_kar': toplam_kar_marka,
                'toplam_gelir': toplam_gelir_marka
            })
        
        marka_karlilik.append({
            'marka': marka,
            'toplam_kar': toplam_kar_marka,
            'ortalama_kar': ortalama_kar_marka,
            'satis_sayisi': satis_sayisi
        })
    
    marka_karlilik = sorted(marka_karlilik, key=lambda x: x['toplam_kar'], reverse=True)
    marka_kar_marji = sorted(marka_kar_marji, key=lambda x: x['kar_marji_yuzdesi'], reverse=True)
    
    # En yüksek karlılık değerini bul (progress bar için)
    max_kar = max([item['toplam_kar'] for item in marka_karlilik], default=1)
    
    # Progress bar için yüzde hesapla
    for item in marka_karlilik:
        if max_kar > 0:
            item['kar_yuzdesi'] = int((item['toplam_kar'] / max_kar) * 100)
        else:
            item['kar_yuzdesi'] = 0
    
    # ============ PERFORMANS METRİKLERİ ============
    # Ortalama satış süresi hesaplama
    satis_suresi_araclar = satilan_araclar.filter(
        satis_tarihi__isnull=False, ilan_tarihi__isnull=False
    )
    
    toplam_gun = 0
    arac_sayisi = 0
    
    for arac in satis_suresi_araclar:
        sure = (arac.satis_tarihi - arac.ilan_tarihi).days
        if sure >= 0:
            toplam_gun += sure
            arac_sayisi += 1
    
    ortalama_satis_suresi = toplam_gun / arac_sayisi if arac_sayisi > 0 else None
    
    # En hızlı satılan araçlar
    en_hizli_satilan = []
    for arac in satis_suresi_araclar:
        sure = (arac.satis_tarihi - arac.ilan_tarihi).days
        if sure >= 0:
            en_hizli_satilan.append({
                'arac': arac,
                'sure': sure
            })
    en_hizli_satilan = sorted(en_hizli_satilan, key=lambda x: x['sure'])[:10]
    
    # En yavaş satılan araçlar
    en_yavas_satilan = sorted(en_hizli_satilan, key=lambda x: x['sure'], reverse=True)[:10] if en_hizli_satilan else []
    
    # ============ ENVANTER ANALİZİ ============
    # Mevcut stok
    mevcut_stok = Arac.objects.filter(satildi_mi=False).count()
    
    # Satılmayan araçlar (stokta kalanlar)
    satilmayan_araclar = Arac.objects.filter(satildi_mi=False).order_by('-ilan_tarihi')
    
    # Stokta kalma süreleri
    stokta_kalanlar = []
    for arac in satilmayan_araclar[:20]:
        sure = (timezone.now() - arac.ilan_tarihi).days
        stokta_kalanlar.append({
            'arac': arac,
            'sure': sure
        })
    
    # Durum bazında satış performansı
    durum_performans = []
    for durum_choice in ['yeni', 'firsat']:
        durum_araclari = satilan_araclar.filter(durum=durum_choice)
        if durum_araclari.exists():
            toplam_gelir_durum = _get_toplam_gelir(durum_araclari)
            toplam_kar_durum = _get_toplam_kar(durum_araclari)
            
            durum_performans.append({
                'durum': durum_choice,
                'satis_sayisi': durum_araclari.count(),
                'toplam_gelir': toplam_gelir_durum,
                'toplam_kar': toplam_kar_durum
            })
    
    # Yıl bazında araç dağılımı
    yil_dagilimi = []
    yillar = satilan_araclar.values_list('yil', flat=True).distinct().order_by('-yil')
    for yil in yillar:
        yil_dagilimi.append({
            'yil': yil,
            'satis_sayisi': satilan_araclar.filter(yil=yil).count()
        })
    
    # ============ MARKA ANALİZLERİ ============
    # Markalara göre toplam satış sayısı
    marka_toplam_satis = []
    for marka in markalar:
        marka_toplam_satis.append({
            'marka': marka,
            'toplam': satilan_araclar.filter(marka__ad=marka).count()
        })
    marka_toplam_satis = sorted(marka_toplam_satis, key=lambda x: x['toplam'], reverse=True)
    
    # Markalara göre toplam gelir
    marka_toplam_gelir = []
    for marka in markalar:
        marka_araclari = satilan_araclar.filter(marka__ad=marka)
        toplam_gelir_marka = _get_toplam_gelir(marka_araclari)
        marka_toplam_gelir.append({
            'marka': marka,
            'toplam_gelir': toplam_gelir_marka
        })
    marka_toplam_gelir = sorted(marka_toplam_gelir, key=lambda x: x['toplam_gelir'], reverse=True)
    
    # ============ ZAMAN BAZLI ANALİZLER ============
    # Yıllara göre satışlar
    yillik_satislar = {}
    yillik_gelir = {}
    yillik_kar = {}
    
    for yil in range(baslangic_yil, bitis_yil + 1):
        yil_araclari = satilan_araclar.filter(
            Q(satis_tarihi__year=yil) | 
            Q(satis_tarihi__isnull=True, ilan_tarihi__year=yil)
        )
        
        yillik_satislar[yil] = yil_araclari.count()
        yillik_gelir[yil] = _get_toplam_gelir(yil_araclari)
        yillik_kar[yil] = _get_toplam_kar(yil_araclari)
    
    # ============ FİYAT ANALİZLERİ ============
    # Fiyat indirimi olan araçlar
    fiyat_indirimi_araclari = satilan_araclar.filter(
        satis_fiyat__isnull=False
    ).exclude(satis_fiyat__gte=F('ilan_fiyat'))
    
    fiyat_indirimi_sayisi = fiyat_indirimi_araclari.count()
    toplam_fiyat_indirimi = sum([
        (arac.ilan_fiyat - arac.satis_fiyat) 
        for arac in fiyat_indirimi_araclari
    ])
    
    ortalama_fiyat_indirimi = toplam_fiyat_indirimi / fiyat_indirimi_sayisi if fiyat_indirimi_sayisi > 0 else 0
    
    # ============ PERSONEL PERFORMANSI ============
    personel_performans = _get_personel_performans(satilan_araclar)
    
    context = {
        # Temel veriler
        'aylik_satislar': aylik_satislar,
        'aylik_gelir': aylik_gelir,
        'aylik_kar': aylik_kar,
        'marka_aylik_satislar': marka_aylik_satislar,
        
        # Toplam istatistikler
        'toplam_satilan': toplam_satilan,
        'toplam_gelir': toplam_gelir,
        'toplam_kar': toplam_kar,
        'ortalama_satis_fiyati': ortalama_satis_fiyati,
        'ortalama_kar': ortalama_kar,
        
        # Karlılık
        'en_karli_araclar': en_karli_araclar,
        'en_az_karli_araclar': en_az_karli_araclar,
        'marka_karlilik': marka_karlilik,
        'max_kar': max_kar,  # Progress bar için
        
        # Performans
        'ortalama_satis_suresi': ortalama_satis_suresi,
        'en_hizli_satilan': en_hizli_satilan,
        'en_yavas_satilan': en_yavas_satilan,
        
        # Envanter
        'mevcut_stok': mevcut_stok,
        'satilmayan_araclar': satilmayan_araclar[:20],
        'stokta_kalanlar': stokta_kalanlar,
        'durum_performans': durum_performans,
        'yil_dagilimi': yil_dagilimi,
        
        # Marka analizleri
        'marka_toplam_satis': marka_toplam_satis,
        'marka_toplam_gelir': marka_toplam_gelir,
        'marka_kar_marji': marka_kar_marji,  # Kar marjı yüzdesi için
        
        # Zaman bazlı
        'yillik_satislar': yillik_satislar,
        'yillik_gelir': yillik_gelir,
        'yillik_kar': yillik_kar,
        
        # Fiyat analizleri
        'fiyat_indirimi_sayisi': fiyat_indirimi_sayisi,
        'toplam_fiyat_indirimi': toplam_fiyat_indirimi,
        'ortalama_fiyat_indirimi': ortalama_fiyat_indirimi,
        
        # Personel performansı
        'personel_performans': personel_performans,
        
        # Filtre parametreleri
        'baslangic_yil': baslangic_yil,
        'bitis_yil': bitis_yil,
        'baslangic_ay': baslangic_ay,
        'bitis_ay': bitis_ay,
        'mevcut_yil': mevcut_yil,
        'yil_listesi': yil_listesi,
    }
    
    return render(request, 'raporlar/raporlar.html', context)


@rapor_required
def rapor_pdf(request):
    """PDF formatında kapsamlı rapor oluştur"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from io import BytesIO
        
        # Tarih aralığı
        baslangic_yil = request.GET.get('baslangic_yil', timezone.now().year)
        bitis_yil = request.GET.get('bitis_yil', timezone.now().year)
        baslangic_ay = request.GET.get('baslangic_ay', 0)
        bitis_ay = request.GET.get('bitis_ay', 0)
        
        try:
            baslangic_yil = int(baslangic_yil)
            bitis_yil = int(bitis_yil)
            baslangic_ay = int(baslangic_ay)
            bitis_ay = int(bitis_ay)
        except (ValueError, TypeError):
            baslangic_yil = timezone.now().year
            bitis_yil = timezone.now().year
            baslangic_ay = 0
            bitis_ay = 0
        
        # PDF oluştur
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        # Başlık
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
            alignment=TA_CENTER,
        )
        title = Paragraph("Kapsamlı Satış Raporu", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Tarih aralığı
        date_style = ParagraphStyle(
            'CustomDate',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#6b7280'),
            alignment=TA_CENTER,
        )
        date_text = Paragraph(f"Tarih Aralığı: {baslangic_yil} - {bitis_yil}", date_style)
        elements.append(date_text)
        elements.append(Spacer(1, 0.3*inch))
        
        # Toplam istatistikler
        satilan_araclar = Arac.objects.filter(satildi_mi=True)
        toplam_satilan = satilan_araclar.count()
        toplam_gelir = _get_toplam_gelir(satilan_araclar)
        toplam_kar = _get_toplam_kar(satilan_araclar)
        mevcut_stok = Arac.objects.filter(satildi_mi=False).count()
        
        stats_data = [
            ['Toplam Satılan Araç', f"{toplam_satilan} adet"],
            ['Toplam Gelir', f"{toplam_gelir:,} TL"],
            ['Toplam Kar', f"{toplam_kar:,} TL"],
            ['Mevcut Stok', f"{mevcut_stok} adet"],
        ]
        
        stats_table = Table(stats_data, colWidths=[4*inch, 3*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fbbf24')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#000000')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Karlılık Analizi
        karlilik_baslik = Paragraph("Marka Bazında Karlılık", styles['Heading2'])
        elements.append(karlilik_baslik)
        elements.append(Spacer(1, 0.2*inch))
        
        marka_karlilik = satilan_araclar.values('marka').annotate(
            toplam_kar=Sum('kar_zarar'),
            ortalama_kar=Avg('kar_zarar'),
            satis_sayisi=Count('id')
        ).order_by('-toplam_kar')
        
        karlilik_data = [['Marka', 'Satış Sayısı', 'Toplam Kar (TL)', 'Ort. Kar (TL)']]
        for item in marka_karlilik:
            if item['toplam_kar']:
                karlilik_data.append([
                    item['marka'],
                    str(item['satis_sayisi']),
                    f"{item['toplam_kar']:,}",
                    f"{item['ortalama_kar']:.0f}"
                ])
        
        if len(karlilik_data) > 1:
            karlilik_table = Table(karlilik_data, colWidths=[2*inch, 1.5*inch, 2*inch, 1.5*inch])
            karlilik_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            elements.append(karlilik_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Aylık satışlar
        aylik_baslik = Paragraph("Aylık Satışlar", styles['Heading2'])
        elements.append(aylik_baslik)
        elements.append(Spacer(1, 0.2*inch))
        
        aylik_data = [['Ay', 'Satış', 'Gelir (TL)', 'Kar (TL)']]
        for yil in range(baslangic_yil, bitis_yil + 1):
            for ay in range(1, 13):
                # Satış tarihi yoksa ilan tarihini kullan
                ay_satislari = satilan_araclar.filter(
                    Q(satis_tarihi__year=yil, satis_tarihi__month=ay) |
                    Q(satis_tarihi__isnull=True, ilan_tarihi__year=yil, ilan_tarihi__month=ay)
                )
                satislar = ay_satislari.count()
                if satislar > 0:
                    gelir = _get_toplam_gelir(ay_satislari)
                    kar = _get_toplam_kar(ay_satislari)
                    ay_adi = f"{month_name[ay]} {yil}"
                    aylik_data.append([ay_adi, str(satislar), f"{gelir:,}", f"{kar:,}"])
        
        if len(aylik_data) > 1:
            aylik_table = Table(aylik_data, colWidths=[2.5*inch, 1.5*inch, 2*inch, 1.5*inch])
            aylik_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            elements.append(aylik_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Markalara göre satışlar
        marka_baslik = Paragraph("Markalara Göre Satışlar", styles['Heading2'])
        elements.append(marka_baslik)
        elements.append(Spacer(1, 0.2*inch))
        
        marka_data = [['Marka', 'Satış Sayısı', 'Toplam Gelir (TL)']]
        marka_satislar = satilan_araclar.values('marka__ad').annotate(
            toplam=Count('id'),
            toplam_gelir=Sum(Case(
                When(satis_fiyat__isnull=False, then=F('satis_fiyat')),
                default=F('ilan_fiyat'),
                output_field=IntegerField()
            ))
        ).order_by('-toplam')
        
        for item in marka_satislar:
            marka_data.append([
                item['marka__ad'],
                str(item['toplam']),
                f"{item['toplam_gelir']:,}"
            ])
        
        if len(marka_data) > 1:
            marka_table = Table(marka_data, colWidths=[2.5*inch, 2.5*inch, 2*inch])
            marka_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#000000')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            elements.append(marka_table)
        
        # PDF'i oluştur
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="kapsamli_satis_raporu_{baslangic_yil}_{bitis_yil}.pdf"'
        return response
        
    except ImportError:
        messages.error(request, 'PDF oluşturmak için reportlab kütüphanesi gerekli. Lütfen yükleyin: pip install reportlab')
        return redirect('raporlar:raporlar')


@rapor_required
def rapor_excel(request):
    """Excel formatında kapsamlı rapor oluştur"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference
        from django.utils import timezone
        from calendar import month_name
        from io import BytesIO
        
        # Tarih aralığı
        baslangic_yil = request.GET.get('baslangic_yil', timezone.now().year)
        bitis_yil = request.GET.get('bitis_yil', timezone.now().year)
        baslangic_ay = request.GET.get('baslangic_ay', 0)
        bitis_ay = request.GET.get('bitis_ay', 0)
        
        try:
            baslangic_yil = int(baslangic_yil)
            bitis_yil = int(bitis_yil)
            baslangic_ay = int(baslangic_ay)
            bitis_ay = int(bitis_ay)
        except (ValueError, TypeError):
            baslangic_yil = timezone.now().year
            bitis_yil = timezone.now().year
            baslangic_ay = 0
            bitis_ay = 0
        
        # Excel çalışma kitabı oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Kapsamlı Rapor"
        
        # Stil tanımlamaları
        title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        cell_font = Font(name='Arial', size=11)
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Başlık
        ws.merge_cells('A1:E1')
        ws['A1'] = "Kapsamlı Satış Raporu"
        ws['A1'].font = title_font
        ws['A1'].fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        ws['A1'].alignment = center_alignment
        ws.row_dimensions[1].height = 30
        
        # Tarih aralığı
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Tarih Aralığı: {baslangic_yil} - {bitis_yil}"
        ws['A2'].alignment = center_alignment
        ws.row_dimensions[2].height = 20
        
        row = 4
        
        # Toplam istatistikler
        satilan_araclar = Arac.objects.filter(satildi_mi=True)
        toplam_satilan = satilan_araclar.count()
        toplam_gelir = _get_toplam_gelir(satilan_araclar)
        toplam_kar = _get_toplam_kar(satilan_araclar)
        mevcut_stok = Arac.objects.filter(satildi_mi=False).count()
        
        ws['A4'] = "Toplam Satılan Araç"
        ws['B4'] = f"{toplam_satilan} adet"
        ws['A5'] = "Toplam Gelir"
        ws['B5'] = f"{toplam_gelir:,} TL"
        ws['A6'] = "Toplam Kar"
        ws['B6'] = f"{toplam_kar:,} TL"
        ws['A7'] = "Mevcut Stok"
        ws['B7'] = f"{mevcut_stok} adet"
        
        for r in range(4, 8):
            ws[f'A{r}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'B{r}'].font = cell_font
        
        row = 9
        
        # Karlılık Analizi
        ws[f'A{row}'] = "Marka Bazında Karlılık"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Marka"
        ws[f'B{row}'] = "Satış Sayısı"
        ws[f'C{row}'] = "Toplam Kar (TL)"
        ws[f'D{row}'] = "Ortalama Kar (TL)"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].alignment = center_alignment
        row += 1
        
        marka_karlilik = satilan_araclar.values('marka').annotate(
            toplam_kar=Sum('kar_zarar'),
            ortalama_kar=Avg('kar_zarar'),
            satis_sayisi=Count('id')
        ).order_by('-toplam_kar')
        
        for item in marka_karlilik:
            if item['toplam_kar']:
                ws[f'A{row}'] = item['marka']
                ws[f'B{row}'] = item['satis_sayisi']
                ws[f'C{row}'] = item['toplam_kar']
                ws[f'D{row}'] = int(item['ortalama_kar'])
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].font = cell_font
                    if col != 'A':
                        ws[f'{col}{row}'].alignment = center_alignment
                row += 1
        
        row += 2
        
        # Aylık satışlar
        ws[f'A{row}'] = "Aylık Satışlar"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Ay"
        ws[f'B{row}'] = "Satış Sayısı"
        ws[f'C{row}'] = "Gelir (TL)"
        ws[f'D{row}'] = "Kar (TL)"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].alignment = center_alignment
        row += 1
        
        for yil in range(baslangic_yil, bitis_yil + 1):
            for ay in range(1, 13):
                # Satış tarihi yoksa ilan tarihini kullan
                ay_satislari = satilan_araclar.filter(
                    Q(satis_tarihi__year=yil, satis_tarihi__month=ay) |
                    Q(satis_tarihi__isnull=True, ilan_tarihi__year=yil, ilan_tarihi__month=ay)
                )
                satislar = ay_satislari.count()
                if satislar > 0:
                    gelir = _get_toplam_gelir(ay_satislari)
                    kar = _get_toplam_kar(ay_satislari)
                    ay_adi = f"{month_name[ay]} {yil}"
                    ws[f'A{row}'] = ay_adi
                    ws[f'B{row}'] = satislar
                    ws[f'C{row}'] = gelir
                    ws[f'D{row}'] = kar
                    ws[f'A{row}'].font = cell_font
                    for col in ['B', 'C', 'D']:
                        ws[f'{col}{row}'].font = cell_font
                        ws[f'{col}{row}'].alignment = center_alignment
                    row += 1
        
        row += 2
        
        # Markalara göre satışlar
        ws[f'A{row}'] = "Markalara Göre Satışlar"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Marka"
        ws[f'B{row}'] = "Satılan Araç Sayısı"
        ws[f'C{row}'] = "Toplam Gelir (TL)"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
            ws[f'{col}{row}'].font = Font(name='Arial', size=12, bold=True, color='000000')
            ws[f'{col}{row}'].alignment = center_alignment
        row += 1
        
        marka_satislar = satilan_araclar.values('marka__ad').annotate(
            toplam=Count('id'),
            toplam_gelir=Sum(Case(
                When(satis_fiyat__isnull=False, then=F('satis_fiyat')),
                default=F('ilan_fiyat'),
                output_field=IntegerField()
            ))
        ).order_by('-toplam')
        
        for item in marka_satislar:
            ws[f'A{row}'] = item['marka__ad']
            ws[f'B{row}'] = item['toplam']
            ws[f'C{row}'] = item['toplam_gelir']
            ws[f'A{row}'].font = cell_font
            ws[f'B{row}'].font = cell_font
            ws[f'C{row}'].font = cell_font
            ws[f'B{row}'].alignment = center_alignment
            ws[f'C{row}'].alignment = center_alignment
            row += 1
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        
        # Excel dosyasını oluştur
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="kapsamli_satis_raporu_{baslangic_yil}_{bitis_yil}.xlsx"'
        return response
        
    except ImportError:
        messages.error(request, 'Excel oluşturmak için openpyxl kütüphanesi gerekli. Lütfen yükleyin: pip install openpyxl')
        return redirect('raporlar:raporlar')
